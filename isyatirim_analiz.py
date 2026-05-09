from __future__ import annotations

import argparse
import concurrent.futures as cf
import datetime as dt
import html as html_lib
import json
import os
import re
import time
from dataclasses import asdict, dataclass
from html.parser import HTMLParser
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple
from urllib.request import Request, urlopen

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MIDAS_URL = "https://www.getmidas.com/temettu-takvim/"
BIST_LIST_URL = "https://stockanalysis.com/list/borsa-istanbul/"
ISYATIRIM_URL = "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse={ticker}"
DEFAULT_OUTPUT_DIR = Path.cwd()
DEFAULT_OUTPUT_NAME = "isyatirim_analiz_raporu.xlsx"
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)
TR_MONTHS = {
    "Ocak": 1,
    "Şubat": 2,
    "Mart": 3,
    "Nisan": 4,
    "Mayıs": 5,
    "Haziran": 6,
    "Temmuz": 7,
    "Ağustos": 8,
    "Eylül": 9,
    "Ekim": 10,
    "Kasım": 11,
    "Aralık": 12,
}
SCRIPT_VERSION = "1.0.0"


@dataclass(frozen=True)
class DividendRow:
    ticker: str
    payment_date: dt.date
    cash_dividend_tl: float
    net_dividend_per_share_tl: float


@dataclass(frozen=True)
class CompanySnapshot:
    ticker: str
    company_name: str
    recommendation: Optional[str]
    recommendation_date: Optional[str]
    target_price_try: Optional[float]
    upside_pct: Optional[float]
    analyst_name: Optional[str]
    current_price_try: Optional[float]
    current_pe: Optional[float]
    current_fd_favok: Optional[float]
    current_fd_sales: Optional[float]
    current_pd_dd: Optional[float]
    current_foreign_ratio_pct: Optional[float]
    latest_sales_mn_try: Optional[float]
    latest_ebitda_mn_try: Optional[float]
    latest_net_income_mn_try: Optional[float]
    previous_sales_mn_try: Optional[float]
    previous_ebitda_mn_try: Optional[float]
    previous_net_income_mn_try: Optional[float]
    next_dividend_yield_pct: Optional[float]
    next_dividend_share_tl: Optional[float]
    next_dividend_total_mn_try: Optional[float]
    next_dividend_date: Optional[str]
    annual_dividend_yield_pct: Optional[float]
    ownership_top_holder: Optional[str]
    ownership_top_ratio_pct: Optional[float]
    sector: Optional[str]
    founding_date: Optional[str]
    activity: Optional[str]
    phone: Optional[str]


@dataclass(frozen=True)
class ScoredCompany:
    ticker: str
    company_name: str
    score: int
    decision: str
    rationale: str
    snapshot: CompanySnapshot
    sales_growth_pct: Optional[float]
    net_income_growth_pct: Optional[float]
    ebitda_growth_pct: Optional[float]


class TableParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tables: List[List[List[str]]] = []
        self._in_table = False
        self._in_row = False
        self._in_cell = False
        self._current_table: List[List[str]] = []
        self._current_row: List[str] = []
        self._cell_parts: List[str] = []

    def handle_starttag(self, tag: str, attrs):
        if tag == "table":
            self._in_table = True
            self._current_table = []
        elif tag == "tr" and self._in_table:
            self._in_row = True
            self._current_row = []
        elif tag in {"td", "th"} and self._in_row:
            self._in_cell = True
            self._cell_parts = []

    def handle_endtag(self, tag: str):
        if tag in {"td", "th"} and self._in_cell:
            text = html_lib.unescape("".join(self._cell_parts)).strip()
            self._current_row.append(re.sub(r"\s+", " ", text))
            self._in_cell = False
        elif tag == "tr" and self._in_row:
            if self._current_row:
                self._current_table.append(self._current_row)
            self._current_row = []
            self._in_row = False
        elif tag == "table" and self._in_table:
            if self._current_table:
                self.tables.append(self._current_table)
            self._current_table = []
            self._in_table = False

    def handle_data(self, data: str):
        if self._in_cell:
            self._cell_parts.append(data)


def parse_number(value: Optional[str]) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s in {"N/A", "None", "-", "Açıklanmadı", "Kayıt bulunamadı."}:
        return None
    s = s.replace("TL", "").replace("mn", "").replace("TRY", "")
    s = s.replace("%", "").replace("\xa0", "").replace(" ", "")
    s = s.replace(".", "")
    if s.count(",") == 1:
        s = s.replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_date_tr(value: str) -> dt.date:
    day, month_name, year = value.split()
    return dt.date(int(year), TR_MONTHS[month_name], int(day))


def fetch_html(url: str, retries: int = 3, timeout: int = 30) -> str:
    req = Request(url, headers={"User-Agent": USER_AGENT, "Accept-Language": "tr-TR,tr;q=0.9,en;q=0.8"})
    last_exc: Optional[Exception] = None
    for attempt in range(retries):
        try:
            with urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", "ignore")
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            time.sleep(min(2**attempt, 4))
    raise RuntimeError(f"Failed to fetch {url}: {last_exc}")


def parse_tables(html_text: str) -> List[List[List[str]]]:
    parser = TableParser()
    parser.feed(html_text)
    return parser.tables


def split_multi_cell(text: str) -> List[str]:
    parts = [p.strip() for p in re.split(r"\s+", text.strip()) if p.strip()]
    return parts


def table_to_dict(table: List[List[str]]) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    for row in table:
        if not row:
            continue
        out[row[0]] = row[1:]
    return out


def first_value(mapping: Dict[str, List[str]], key: str) -> Optional[str]:
    values = mapping.get(key)
    if not values:
        return None
    return values[0]


def parse_company_name(html_text: str, ticker: str) -> str:
    title_match = re.search(r"<title>(.*?)</title>", html_text, re.S | re.I)
    if title_match:
        title = re.sub(r"\s+", " ", html_lib.unescape(title_match.group(1)).strip())
        m = re.search(rf"^{re.escape(ticker)}\s*-\s*(.*?)\s+Hisse Senedi", title, re.I)
        if m:
            return m.group(1).strip()
        if "Hisse Senedi" in title and " - " in title:
            return title.split(" - ", 1)[1].split(" Hisse Senedi", 1)[0].strip()
    h1_match = re.search(r"<h1[^>]*>(.*?)</h1>", html_text, re.S | re.I)
    if h1_match:
        text = re.sub(r"\s+", " ", html_lib.unescape(h1_match.group(1)).strip())
        return text.split("|", 1)[0].strip() or ticker
    return ticker


def extract_recommendation(html_text: str) -> Tuple[Optional[str], Optional[str], Optional[float], Optional[float], Optional[str]]:
    def grab(pattern: str) -> Optional[str]:
        match = re.search(pattern, html_text, re.S | re.I)
        return match.group(1).strip() if match else None

    label = grab(r'id="Oneri_Aciklama"[^>]*>([^<]+)</span>')
    date = grab(r'Son Öneri Tarihi\s*<span>([^<]+)</span>')
    target = parse_number(grab(r'Hedef Fiyat\s*<span>([^<]+)</span>'))
    upside = parse_number(grab(r'Getiri Pot\s*<span>%?([^<]+)</span>'))
    analyst = grab(r'<h4>([^<]+)</h4>')
    return label, date, target, upside, analyst


def _find_table(tables: List[List[List[str]]], predicate) -> List[List[str]]:
    for table in tables:
        if table and predicate(table):
            return table
    return []


def _table_first_cell(table: List[List[str]]) -> str:
    return table[0][0].strip() if table and table[0] and table[0][0] else ""


def parse_company_page(html_text: str, ticker: str) -> CompanySnapshot:
    tables = parse_tables(html_text)
    company_name = parse_company_name(html_text, ticker)
    recommendation, recommendation_date, target_price_try, upside_pct, analyst_name = extract_recommendation(html_text)

    company_info_table = _find_table(tables, lambda t: _table_first_cell(t) == "Ünvanı")
    ratios_table = _find_table(tables, lambda t: _table_first_cell(t) == "F/K")
    current_ratios_table = _find_table(
        tables,
        lambda t: _table_first_cell(t) == "Kod" and any("Kapanış" in cell for cell in t[0]) and any("F/K" in cell for cell in t[0]),
    )
    financial_tables = [t for t in tables if _table_first_cell(t) == "Satışlar"]
    dividend_table = _find_table(tables, lambda t: _table_first_cell(t) == "Kod" and any("Dağ. Tarihi" in cell for cell in t[0]))

    company_info = table_to_dict(company_info_table) if company_info_table else {}
    current_ratios = table_to_dict(current_ratios_table) if current_ratios_table else {}
    ratios_current = table_to_dict(ratios_table) if ratios_table else {}

    latest_sales_mn_try = latest_ebitda_mn_try = latest_net_income_mn_try = None
    previous_sales_mn_try = previous_ebitda_mn_try = previous_net_income_mn_try = None
    if financial_tables:
        latest = table_to_dict(financial_tables[0])
        latest_sales_mn_try = parse_number(first_value(latest, "Satışlar"))
        latest_ebitda_mn_try = parse_number(first_value(latest, "FAVÖK"))
        latest_net_income_mn_try = parse_number(first_value(latest, "Net Kar"))
    if len(financial_tables) > 1:
        previous = table_to_dict(financial_tables[1])
        previous_sales_mn_try = parse_number(first_value(previous, "Satışlar"))
        previous_ebitda_mn_try = parse_number(first_value(previous, "FAVÖK"))
        previous_net_income_mn_try = parse_number(first_value(previous, "Net Kar"))

    current_price_try = parse_number(first_value(current_ratios, "Kapanış (TL)") or first_value(current_ratios, "Kapanış Fiyat") or first_value(current_ratios, "Kapanış Fiyat (TL)"))
    current_pe = parse_number(first_value(current_ratios, "F/K")) or parse_number(first_value(ratios_current, "F/K"))
    current_fd_favok = parse_number(first_value(current_ratios, "FD/FAVÖK")) or parse_number(first_value(ratios_current, "FD/FAVÖK"))
    current_fd_sales = parse_number(first_value(current_ratios, "FD/Satışlar")) or parse_number(first_value(ratios_current, "FD/Satışlar"))
    current_pd_dd = parse_number(first_value(current_ratios, "PD/DD")) or parse_number(first_value(ratios_current, "PD/DD"))
    current_foreign_ratio_pct = parse_number(first_value(ratios_current, "Yabancı Oranı (%)"))

    next_dividend_yield_pct = next_dividend_share_tl = next_dividend_total_mn_try = None
    next_dividend_date = None
    if dividend_table and len(dividend_table) > 1:
        row = dividend_table[1]
        if len(row) >= 3:
            next_dividend_date = row[1] if len(row) > 1 else None
            next_dividend_yield_pct = parse_number(row[2]) if len(row) > 2 else None
            next_dividend_share_tl = parse_number(row[3]) if len(row) > 3 else None
            next_dividend_total_mn_try = parse_number(row[5]) if len(row) > 5 else None

    annual_dividend_yield_pct = None
    dividend_summary_table = _find_table(tables, lambda t: _table_first_cell(t) == "Kod" and any("Temettü Verim" in cell for cell in t[0]) and any("Tahmin/Gerçekleşen" in cell for cell in t[0]))
    if dividend_summary_table and len(dividend_summary_table) > 1:
        vals = split_multi_cell(dividend_summary_table[1][2]) if len(dividend_summary_table[1]) > 2 else []
        annual_dividend_yield_pct = parse_number(vals[0]) if vals else None

    return CompanySnapshot(
        ticker=ticker,
        company_name=company_name,
        recommendation=recommendation,
        recommendation_date=recommendation_date,
        target_price_try=target_price_try,
        upside_pct=upside_pct,
        analyst_name=analyst_name,
        current_price_try=current_price_try,
        current_pe=current_pe,
        current_fd_favok=current_fd_favok,
        current_fd_sales=current_fd_sales,
        current_pd_dd=current_pd_dd,
        current_foreign_ratio_pct=current_foreign_ratio_pct,
        latest_sales_mn_try=latest_sales_mn_try,
        latest_ebitda_mn_try=latest_ebitda_mn_try,
        latest_net_income_mn_try=latest_net_income_mn_try,
        previous_sales_mn_try=previous_sales_mn_try,
        previous_ebitda_mn_try=previous_ebitda_mn_try,
        previous_net_income_mn_try=previous_net_income_mn_try,
        next_dividend_yield_pct=next_dividend_yield_pct,
        next_dividend_share_tl=next_dividend_share_tl,
        next_dividend_total_mn_try=next_dividend_total_mn_try,
        next_dividend_date=next_dividend_date,
        annual_dividend_yield_pct=annual_dividend_yield_pct,
        ownership_top_holder=None,
        ownership_top_ratio_pct=None,
        sector=first_value(company_info, "Faal Alanı"),
        founding_date=first_value(company_info, "Kuruluş"),
        activity=first_value(company_info, "Faal Alanı"),
        phone=first_value(company_info, "Telefon"),
    )


def growth_pct(current: Optional[float], previous: Optional[float]) -> Optional[float]:
    if current is None or previous in (None, 0):
        return None
    return ((current - previous) / abs(previous)) * 100


def score_company(snapshot: CompanySnapshot) -> ScoredCompany:
    score = 0
    reasons: List[str] = []

    rec = (snapshot.recommendation or "").upper()
    if rec == "AL":
        score += 25
        reasons.append("İş Yatırım önerisi AL")
    elif rec == "TUT":
        score += 8
        reasons.append("İş Yatırım önerisi TUT")
    elif rec == "SAT":
        score -= 25
        reasons.append("İş Yatırım önerisi SAT")

    if snapshot.upside_pct is not None:
        upside_score = max(min(snapshot.upside_pct, 40.0), -20.0)
        score += int(round(upside_score * 0.5))
        reasons.append(f"Getiri potansiyeli %{snapshot.upside_pct:.1f}")

    if snapshot.current_pe is not None:
        if snapshot.current_pe <= 10:
            score += 15
        elif snapshot.current_pe <= 15:
            score += 12
        elif snapshot.current_pe <= 20:
            score += 8
        elif snapshot.current_pe <= 30:
            score += 4
        else:
            score -= 4
        reasons.append(f"F/K {snapshot.current_pe:.1f}")

    if snapshot.current_fd_favok is not None:
        if snapshot.current_fd_favok <= 5:
            score += 15
        elif snapshot.current_fd_favok <= 8:
            score += 12
        elif snapshot.current_fd_favok <= 12:
            score += 8
        elif snapshot.current_fd_favok <= 16:
            score += 4
        else:
            score -= 3
        reasons.append(f"FD/FAVÖK {snapshot.current_fd_favok:.1f}")

    if snapshot.current_pd_dd is not None:
        if snapshot.current_pd_dd <= 1.5:
            score += 10
        elif snapshot.current_pd_dd <= 2.5:
            score += 8
        elif snapshot.current_pd_dd <= 3.5:
            score += 4
        else:
            score -= 2
        reasons.append(f"PD/DD {snapshot.current_pd_dd:.1f}")

    sales_growth = growth_pct(snapshot.latest_sales_mn_try, snapshot.previous_sales_mn_try)
    net_income_growth = growth_pct(snapshot.latest_net_income_mn_try, snapshot.previous_net_income_mn_try)
    ebitda_growth = growth_pct(snapshot.latest_ebitda_mn_try, snapshot.previous_ebitda_mn_try)

    if sales_growth is not None:
        if sales_growth > 15:
            score += 10
        elif sales_growth > 8:
            score += 7
        elif sales_growth > 0:
            score += 3
        else:
            score -= 3
        reasons.append(f"Satış büyümesi %{sales_growth:.1f}")

    if net_income_growth is not None:
        if net_income_growth > 20:
            score += 10
        elif net_income_growth > 10:
            score += 7
        elif net_income_growth > 0:
            score += 3
        else:
            score -= 5
        reasons.append(f"Net kâr büyümesi %{net_income_growth:.1f}")

    if ebitda_growth is not None:
        if ebitda_growth > 15:
            score += 6
        elif ebitda_growth > 5:
            score += 4
        elif ebitda_growth > 0:
            score += 2
        else:
            score -= 2
        reasons.append(f"FAVÖK büyümesi %{ebitda_growth:.1f}")

    if snapshot.next_dividend_yield_pct is not None:
        if snapshot.next_dividend_yield_pct > 2:
            score += 8
        elif snapshot.next_dividend_yield_pct > 1:
            score += 5
        elif snapshot.next_dividend_yield_pct > 0:
            score += 2
        reasons.append(f"Yaklaşan temettü verimi %{snapshot.next_dividend_yield_pct:.2f}")

    if snapshot.current_foreign_ratio_pct is not None:
        if snapshot.current_foreign_ratio_pct >= 40:
            score += 2
        reasons.append(f"Yabancı oranı %{snapshot.current_foreign_ratio_pct:.2f}")

    score = max(0, min(100, score))
    if score >= 65:
        decision = "BUY"
    elif score >= 45:
        decision = "WATCH"
    else:
        decision = "AVOID"

    return ScoredCompany(
        ticker=snapshot.ticker,
        company_name=snapshot.company_name,
        score=score,
        decision=decision,
        rationale="; ".join(reasons),
        snapshot=snapshot,
        sales_growth_pct=sales_growth,
        net_income_growth_pct=net_income_growth,
        ebitda_growth_pct=ebitda_growth,
    )


def parse_midas(html_text: str) -> List[DividendRow]:
    pattern = re.compile(
        r'<tr class="table-row">\s*'
        r'<td class="code val"><a href="/canli-borsa/[^\"]+-hisse/">(?P<ticker>[A-Z0-9]+)</a></td>\s*'
        r'<td class="val">(?P<date>[^<]+)</td>\s*'
        r'<td class="val">(?P<cash>[^<]+)</td>\s*'
        r'<td class="val">(?P<net>[^<]+)</td>\s*'
        r'</tr>',
        re.S,
    )
    rows: List[DividendRow] = []
    for match in pattern.finditer(html_text):
        rows.append(
            DividendRow(
                ticker=match.group("ticker").strip(),
                payment_date=parse_date_tr(match.group("date").strip()),
                cash_dividend_tl=parse_number(match.group("cash")) or 0.0,
                net_dividend_per_share_tl=parse_number(match.group("net")) or 0.0,
            )
        )
    return rows


def fetch_bist_tickers() -> List[str]:
    html_text = fetch_html(BIST_LIST_URL)
    tickers = re.findall(r'<a href="/quote/ist/([A-Z0-9\.]+)/">\1</a>', html_text)
    return list(dict.fromkeys(tickers))


def fetch_workflow(
    max_workers: int = 6,
    tickers: Optional[Sequence[str]] = None,
    universe: str = "midas",
) -> Tuple[List[DividendRow], List[CompanySnapshot]]:
    if universe == "bist":
        midas_rows: List[DividendRow] = []
        tickers_list = fetch_bist_tickers()
        if tickers:
            wanted = {t.upper().strip() for t in tickers if t.strip()}
            tickers_list = [ticker for ticker in tickers_list if ticker in wanted]
    else:
        midas_rows = parse_midas(fetch_html(MIDAS_URL))
        if tickers:
            wanted = {t.upper().strip() for t in tickers if t.strip()}
            midas_rows = [row for row in midas_rows if row.ticker in wanted]
        tickers_list = sorted({row.ticker for row in midas_rows})

    def load_ticker(ticker: str) -> CompanySnapshot:
        html_text = fetch_html(ISYATIRIM_URL.format(ticker=ticker))
        return parse_company_page(html_text, ticker)

    snapshots: List[CompanySnapshot] = []
    with cf.ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(load_ticker, ticker): ticker for ticker in tickers_list}
        for future in cf.as_completed(futures):
            ticker = futures[future]
            try:
                snapshots.append(future.result())
            except Exception as exc:  # noqa: BLE001
                print(f"[warn] İş Yatırım fetch failed for {ticker}: {exc}", flush=True)
    return midas_rows, snapshots


def combine_rows(midas_rows: List[DividendRow], snapshots: List[CompanySnapshot]) -> List[ScoredCompany]:
    scored = [score_company(snapshot) for snapshot in snapshots]
    if midas_rows:
        by_ticker = {row.ticker: row for row in midas_rows}
        combined = [item for item in scored if item.ticker in by_ticker]
    else:
        combined = scored
    combined.sort(key=lambda x: (x.score, x.snapshot.upside_pct or -1, x.ticker), reverse=True)
    return combined


def build_caption(rows: List[ScoredCompany]) -> str:
    top5 = [row for row in rows if row.decision == "BUY"][:5]
    if not top5:
        top5 = rows[:5]
    pieces = ["İş Yatırım temelli hisse listesi hazır.", "", "Öne çıkanlar:"]
    for idx, row in enumerate(top5, 1):
        upside = f"%{row.snapshot.upside_pct:.1f}" if row.snapshot.upside_pct is not None else "n/a"
        pieces.append(f"{idx}. {row.ticker} — skor {row.score}/100 — {upside} — {row.decision}")
    return "\n".join(pieces)


def default_output_path(output: Optional[str], universe: str = "midas") -> Path:
    if output:
        return Path(output).expanduser().resolve()
    stamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    prefix = "bist_genel_analiz_raporu" if universe == "bist" else "isyatirim_analiz_raporu"
    return (DEFAULT_OUTPUT_DIR / f"{prefix}_{stamp}.xlsx").resolve()


def write_excel(rows: List[ScoredCompany], output_path: Path, universe: str = "midas") -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"
    headers = [
        "Ticker",
        "Company Name",
        "Decision",
        "Score",
        "Recommendation",
        "Recommendation Date",
        "Target Price (TL)",
        "Upside %",
        "Current Price (TL)",
        "F/K",
        "FD/FAVÖK",
        "PD/DD",
        "Latest Sales (mn TL)",
        "Previous Sales (mn TL)",
        "Sales Growth %",
        "Latest Net Income (mn TL)",
        "Previous Net Income (mn TL)",
        "Net Income Growth %",
        "Latest EBITDA (mn TL)",
        "Previous EBITDA (mn TL)",
        "EBITDA Growth %",
        "Next Dividend Yield %",
        "Next Dividend Share (TL)",
        "Next Dividend Date",
        "Rationale",
    ]
    ws.append(headers)
    for row in rows:
        s = row.snapshot
        ws.append([
            row.ticker,
            row.company_name,
            row.decision,
            row.score,
            s.recommendation,
            s.recommendation_date,
            s.target_price_try,
            s.upside_pct,
            s.current_price_try,
            s.current_pe,
            s.current_fd_favok,
            s.current_pd_dd,
            s.latest_sales_mn_try,
            s.previous_sales_mn_try,
            row.sales_growth_pct,
            s.latest_net_income_mn_try,
            s.previous_net_income_mn_try,
            row.net_income_growth_pct,
            s.latest_ebitda_mn_try,
            s.previous_ebitda_mn_try,
            row.ebitda_growth_pct,
            s.next_dividend_yield_pct,
            s.next_dividend_share_tl,
            s.next_dividend_date,
            row.rationale,
        ])

    ws_buy = wb.create_sheet("Buy_List")
    ws_buy.append(headers)
    for row in [r for r in rows if r.decision == "BUY"]:
        s = row.snapshot
        ws_buy.append([
            row.ticker,
            row.company_name,
            row.decision,
            row.score,
            s.recommendation,
            s.recommendation_date,
            s.target_price_try,
            s.upside_pct,
            s.current_price_try,
            s.current_pe,
            s.current_fd_favok,
            s.current_pd_dd,
            s.latest_sales_mn_try,
            s.previous_sales_mn_try,
            row.sales_growth_pct,
            s.latest_net_income_mn_try,
            s.previous_net_income_mn_try,
            row.net_income_growth_pct,
            s.latest_ebitda_mn_try,
            s.previous_ebitda_mn_try,
            row.ebitda_growth_pct,
            s.next_dividend_yield_pct,
            s.next_dividend_share_tl,
            s.next_dividend_date,
            row.rationale,
        ])

    ws_raw = wb.create_sheet("Raw_Companies")
    ws_raw.append(["Ticker", "Company Name", "Raw JSON"])
    for row in rows:
        ws_raw.append([row.ticker, row.company_name, json.dumps(asdict(row.snapshot), ensure_ascii=False)])

    ws_notes = wb.create_sheet("Notes")
    ws_notes.append(["Field", "Value"])
    notes = [
        ("Script version", SCRIPT_VERSION),
        ("Run date", dt.datetime.now().isoformat(timespec="seconds")),
        ("Universe", "BIST listed companies" if universe == "bist" else "Midas dividend calendar"),
        ("Universe source", BIST_LIST_URL if universe == "bist" else MIDAS_URL),
        ("Company page source", "https://www.isyatirim.com.tr/tr-tr/analiz/hisse/Sayfalar/sirket-karti.aspx?hisse=<TICKER>"),
        ("Decision rule", "BUY if score >= 65; WATCH if >= 45; otherwise AVOID."),
        ("Score factors", "Recommendation, upside, F/K, FD/FAVÖK, PD/DD, growth, and upcoming dividend yield."),
    ]
    for k, v in notes:
        ws_notes.append([k, v])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in [ws, ws_buy, ws_raw, ws_notes]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is None:
                    continue
                text = cell.value.strftime("%Y-%m-%d") if hasattr(cell.value, "strftime") else str(cell.value)
                max_len = max(max_len, len(text))
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 48)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_json(rows: List[ScoredCompany], output_path: Path) -> None:
    payload = {
        "script_version": SCRIPT_VERSION,
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "rows": [
            {
                "ticker": row.ticker,
                "company_name": row.company_name,
                "decision": row.decision,
                "score": row.score,
                "rationale": row.rationale,
                "snapshot": asdict(row.snapshot),
                "sales_growth_pct": row.sales_growth_pct,
                "net_income_growth_pct": row.net_income_growth_pct,
                "ebitda_growth_pct": row.ebitda_growth_pct,
            }
            for row in rows
        ],
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="İş Yatırım dividend + analysis workflow")
    parser.add_argument("--output", default=None, help="Output Excel file path; defaults to a timestamped filename")
    parser.add_argument("--telegram", action="store_true", help="Send the Excel file to Telegram if bot credentials are set")
    parser.add_argument("--max-workers", type=int, default=6, help="Parallel company-page fetch concurrency")
    parser.add_argument("--tickers", default=None, help="Comma-separated ticker filter (optional)")
    parser.add_argument("--universe", choices=["midas", "bist"], default="midas", help="Data universe to analyze")
    args = parser.parse_args(argv)

    tickers = [t.strip() for t in args.tickers.split(",")] if args.tickers else None
    output_path = default_output_path(args.output, universe=args.universe)
    midas_rows, snapshots = fetch_workflow(max_workers=args.max_workers, tickers=tickers, universe=args.universe)
    rows = combine_rows(midas_rows, snapshots)
    write_excel(rows, output_path, universe=args.universe)
    write_json(rows, output_path.with_suffix(".json"))

    summary = {
        "output": str(output_path),
        "json": str(output_path.with_suffix(".json")),
        "rows": len(rows),
        "buy_count": sum(1 for row in rows if row.decision == "BUY"),
        "top10": [
            {"ticker": row.ticker, "score": row.score, "decision": row.decision}
            for row in rows[:10]
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.telegram:
        token = os.getenv("TELEGRAM_BOT_TOKEN") or os.getenv("BOT_TOKEN")
        chat_id = os.getenv("TELEGRAM_CHAT_ID") or os.getenv("CHAT_ID")
        if not token or not chat_id:
            raise RuntimeError("Telegram env vars missing: TELEGRAM_BOT_TOKEN/BOT_TOKEN and TELEGRAM_CHAT_ID/CHAT_ID")
        url = f"https://api.telegram.org/bot{token}/sendDocument"
        with output_path.open("rb") as fh:
            resp = requests.post(url, data={"chat_id": chat_id, "caption": build_caption(rows)}, files={"document": (output_path.name, fh)}, timeout=60)
        resp.raise_for_status()
        payload = resp.json()
        if not payload.get("ok"):
            raise RuntimeError(f"Telegram send failed: {payload}")
        print("telegram: sent")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
